#!/usr/bin/env python3
"""
Dify Auto Code Generation Script
- Reads dify_results.xlsx to get method_llm (→ paper_method) and file_path (→ PDF name)
- Reads medbench_eval.py as reference_code
- Calls Dify code-generation workflow with paper_method + reference_code text inputs
- Saves generated code to a new .py file named after the source PDF
- Appends results to an output Excel file

Usage:
    python dify_auto_code.py
    python dify_auto_code.py --input dify_results.xlsx --ref_code medbench_eval.py
    python dify_auto_code.py --output code_results.xlsx --code_dir generated_code
"""

import argparse
import json
import sys
import os
import re
import datetime
import requests
import openpyxl

# ── Configuration ────────────────────────────────────────────────────────────
BASE_URL            = "http://YOUR_DIFY_HOST/v1"   # Replace with your Dify server address
API_KEY             = "your-dify-api-key"               # Replace with your Dify application API key
USER_ID             = "abc-123"
DEFAULT_INPUT_XLSX  = "dify_results.xlsx"
DEFAULT_REF_CODE    = "medbench_eval.py"
DEFAULT_OUTPUT_XLSX = "dify_auto_code_results.xlsx"
DEFAULT_CODE_DIR    = "./"
# ─────────────────────────────────────────────────────────────────────────────

HEADERS_AUTH = {"Authorization": f"Bearer {API_KEY}"}


def run_workflow_text(paper_method: str, reference_code: str, debug: bool = False) -> dict:
    """Run the Dify code-generation workflow with text inputs; return final outputs."""
    print(f"\nRunning workflow (streaming) …\n{'─'*60}")
    payload = {
        "inputs": {
            "paper_method":   paper_method,
            "reference_code": reference_code,
        },
        "response_mode": "streaming",
        "user":          USER_ID,
    }
    if debug:
        snippet = paper_method[:300] + "…" if len(paper_method) > 300 else paper_method
        print(f"[debug] POST {BASE_URL}/workflows/run")
        print(f"[debug] paper_method (first 300 chars):\n{snippet}\n")

    final_outputs: dict = {}
    run_id = None

    with requests.post(
        f"{BASE_URL}/workflows/run",
        headers={**HEADERS_AUTH, "Content-Type": "application/json"},
        json=payload,
        stream=True,
        timeout=600,
    ) as resp:
        if not resp.ok:
            body = resp.content.decode("utf-8", errors="replace")
            raise requests.HTTPError(
                f"{resp.status_code} {resp.reason} for url: {resp.url}\nBody: {body}",
                response=resp,
            )

        for raw_line in resp.iter_lines():
            if not raw_line:
                continue
            line = raw_line.decode("utf-8") if isinstance(raw_line, bytes) else raw_line
            if line.startswith("data: "):
                line = line[len("data: "):]
            try:
                event = json.loads(line)
            except json.JSONDecodeError:
                print(line)
                continue

            event_type = event.get("event", "")

            if event_type == "workflow_started":
                run_id = event.get("workflow_run_id")
                print(f"[workflow_started]  run_id={run_id}")

            elif event_type == "node_started":
                data = event.get("data", {})
                print(f"  [node_started]   {data.get('title', data.get('node_id', ''))}")

            elif event_type == "node_finished":
                data   = event.get("data", {})
                title  = data.get("title", data.get("node_id", ""))
                status = data.get("status", "")
                print(f"  [node_finished]  {title}  status={status}")

            elif event_type == "workflow_finished":
                data          = event.get("data", {})
                final_outputs = data.get("outputs") or {}
                status        = data.get("status", "")
                print(f"\n[workflow_finished] status={status}")
                print(f"  outputs keys: {list(final_outputs.keys())}")

            elif event_type == "text_chunk":
                chunk = event.get("data", {}).get("text", "")
                print(chunk, end="", flush=True)

            elif event_type == "error":
                print(f"[ERROR] {event}", file=sys.stderr)

            else:
                print(f"[{event_type}] {json.dumps(event, ensure_ascii=False)}")

    print(f"\n{'─'*60}\nDone.")
    return {"run_id": run_id, "outputs": final_outputs}


def _clean(value):
    """Remove characters illegal in Excel cells (ASCII control chars except tab/LF/CR)."""
    if not isinstance(value, str):
        return value
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)


def save_to_excel(source_pdf: str, result: dict, output_xlsx: str) -> None:
    """Append a row (source_pdf + workflow outputs) to an Excel file."""
    outputs = result.get("outputs", {})
    run_id  = result.get("run_id", "")
    now     = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(output_xlsx):
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        header = ["timestamp", "source_pdf", "run_id"] + [str(k) for k in outputs.keys()]
        ws.append(header)

    header_row = [cell.value for cell in ws[1]]

    for key in outputs.keys():
        col_name = str(key)
        if col_name not in header_row:
            ws.cell(row=1, column=len(header_row) + 1, value=col_name)
            header_row.append(col_name)

    row_data = [""] * len(header_row)
    fixed = {"timestamp": now, "source_pdf": source_pdf, "run_id": run_id}
    for key, val in fixed.items():
        if key in header_row:
            row_data[header_row.index(key)] = _clean(str(val))

    for key, val in outputs.items():
        col_name = str(key)
        if col_name in header_row:
            cell_val = (
                val if isinstance(val, (str, int, float, bool)) or val is None
                else json.dumps(val, ensure_ascii=False)
            )
            row_data[header_row.index(col_name)] = _clean(str(cell_val) if cell_val is not None else "")

    ws.append(row_data)
    wb.save(output_xlsx)
    print(f"[Excel] Row saved → {output_xlsx}")


def save_code_file(code_content: str, pdf_name: str, code_dir: str) -> str:
    """Save generated code to a .py file named after the source PDF."""
    os.makedirs(code_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(pdf_name))[0]
    safe_name = re.sub(r"[^\w\-.]", "_", base_name)
    code_path = os.path.join(code_dir, f"{safe_name}_medbench_eval.py")
    with open(code_path, "w", encoding="utf-8") as f:
        f.write(code_content)
    print(f"[Code]  Saved → {code_path}")
    return code_path


def read_excel_rows(xlsx_path: str) -> list:
    """Read dify_results.xlsx and return list of dicts, one per data row."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def strip_markdown_fences(text: str) -> str:
    """Remove surrounding markdown code fences (```python ... ``` or ``` ... ```)."""
    text = text.strip()
    text = re.sub(r"^```[a-zA-Z]*\n?", "", text)
    text = re.sub(r"\n?```\s*$", "", text)
    return text.strip()


def extract_code_from_outputs(outputs: dict) -> str:
    """Extract the generated code string from workflow outputs."""
    for key in ("code", "result", "text", "output", "generated_code"):
        val = outputs.get(key)
        if val:
            raw = val if isinstance(val, str) else json.dumps(val, ensure_ascii=False, indent=2)
            return strip_markdown_fences(raw)
    for val in outputs.values():
        if val:
            raw = val if isinstance(val, str) else json.dumps(val, ensure_ascii=False, indent=2)
            return strip_markdown_fences(raw)
    return ""


def main():
    parser = argparse.ArgumentParser(
        description="Generate medbench eval code via Dify for each row in dify_results.xlsx."
    )
    parser.add_argument("--input",    "-i", default=DEFAULT_INPUT_XLSX,
                        help=f"Input Excel with method_llm column (default: {DEFAULT_INPUT_XLSX})")
    parser.add_argument("--ref_code", "-r", default=DEFAULT_REF_CODE,
                        help=f"Reference code file passed as reference_code (default: {DEFAULT_REF_CODE})")
    parser.add_argument("--output",   "-o", default=DEFAULT_OUTPUT_XLSX,
                        help=f"Output Excel for generated-code results (default: {DEFAULT_OUTPUT_XLSX})")
    parser.add_argument("--code_dir", "-c", default=DEFAULT_CODE_DIR,
                        help=f"Directory to save generated .py files (default: {DEFAULT_CODE_DIR})")
    parser.add_argument("--debug",    "-d", action="store_true",
                        help="Print request details before sending")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"[Error] Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)
    if not os.path.exists(args.ref_code):
        print(f"[Error] Reference code file not found: {args.ref_code}", file=sys.stderr)
        sys.exit(1)

    with open(args.ref_code, "r", encoding="utf-8") as f:
        reference_code = f.read()
    print(f"[Init]  Loaded reference code from {args.ref_code} ({len(reference_code)} chars)")

    rows = read_excel_rows(args.input)
    print(f"[Init]  Found {len(rows)} row(s) in {args.input}")

    for i, row in enumerate(rows, 1):
        file_path    = row.get("file_path") or ""
        paper_method = row.get("method_llm") or ""
        pdf_name     = os.path.basename(file_path) if file_path else f"row_{i}.pdf"

        print(f"\n{'═'*60}")
        print(f"[Row {i}/{len(rows)}]  PDF: {pdf_name}")

        if not paper_method:
            print(f"  Skipping — method_llm is empty for this row.")
            continue

        try:
            result       = run_workflow_text(paper_method, reference_code, debug=args.debug)
            outputs      = result.get("outputs", {})
            code_content = extract_code_from_outputs(outputs)

            if code_content:
                save_code_file(code_content, pdf_name, args.code_dir)
            else:
                print(f"  [Warning] No code content found in workflow outputs: {list(outputs.keys())}")

            save_to_excel(pdf_name, result, args.output)

        except requests.HTTPError as e:
            print(f"[HTTP Error] {e}", file=sys.stderr)
        except Exception as e:
            print(f"[Error] {e}", file=sys.stderr)

    print(f"\n{'═'*60}")
    print(f"All rows processed. Results saved to {args.output}")
    print(f"Generated code files saved in: {args.code_dir}/")


if __name__ == "__main__":
    main()
