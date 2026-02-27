#!/usr/bin/env python3
"""
Dify Workflow Script
- Uploads a local file to Dify, then runs the workflow with that file.
- Saves results (file path + workflow outputs) to an Excel file.

Usage:
    python dify_exactor_method.py [--file /path/to/file.pdf] [--output results.xlsx]
"""

import argparse
import json
import sys
import os
import datetime
import requests
import openpyxl

# ── Configuration ────────────────────────────────────────────────────────────
BASE_URL   = "http://YOUR_DIFY_HOST/v1"   # Replace with your Dify server address
API_KEY    = "your-dify-api-key"               # Replace with your Dify application API key
USER_ID    = "abc-123"
# Variable name used in the Dify workflow for the file input
FILE_VAR   = "file"
DEFAULT_FILE   = "example.pdf"
DEFAULT_OUTPUT = "dify_results.xlsx"
# ─────────────────────────────────────────────────────────────────────────────

HEADERS_AUTH = {"Authorization": f"Bearer {API_KEY}"}


def upload_file(file_path: str) -> dict:
    """Upload a local file to Dify and return the response JSON."""
    print(f"[1/2] Uploading file: {file_path}")
    with open(file_path, "rb") as f:
        import mimetypes
        mime_type, _ = mimetypes.guess_type(file_path)
        mime_type = mime_type or "application/octet-stream"
        files = {
            "file": (file_path.split("/")[-1], f, mime_type),
        }
        data = {"user": USER_ID}
        resp = requests.post(
            f"{BASE_URL}/files/upload",
            headers=HEADERS_AUTH,
            files=files,
            data=data,
        )
    resp.raise_for_status()
    result = resp.json()
    print(f"    upload_file_id : {result.get('id')}")
    print(f"    name           : {result.get('name')}")
    return result


def run_workflow(
    upload_file_id: str,
    doc_type: str = "document",
    file_var: str = FILE_VAR,
    debug: bool = False,
) -> dict:
    """Run the Dify workflow in streaming mode and return the final outputs."""
    print(f"\n[2/2] Running workflow (streaming) …\n{'─'*60}")
    payload = {
        "inputs": {
            file_var: {
                "transfer_method": "local_file",
                "upload_file_id": upload_file_id,
                "type": doc_type,
            }
        },
        "response_mode": "streaming",
        "user": USER_ID,
    }
    if debug:
        print(f"[debug] POST {BASE_URL}/workflows/run")
        print(f"[debug] payload:\n{json.dumps(payload, ensure_ascii=False, indent=2)}\n")

    final_outputs = {}
    run_id = None

    with requests.post(
        f"{BASE_URL}/workflows/run",
        headers={**HEADERS_AUTH, "Content-Type": "application/json"},
        json=payload,
        stream=True,
        timeout=300,
    ) as resp:
        if not resp.ok:
            body = resp.content.decode("utf-8", errors="replace")
            raise requests.HTTPError(
                f"{resp.status_code} {resp.reason} for url: {resp.url}\nBody: {body}",
                response=resp,
            )
        for raw_line in resp.iter_lines():
            if not raw_line:
                continue
            line = raw_line.decode("utf-8") if isinstance(raw_line, bytes) else raw_line
            if line.startswith("data: "):
                line = line[len("data: "):]
            try:
                event = json.loads(line)
            except json.JSONDecodeError:
                print(line)
                continue

            event_type = event.get("event", "")

            if event_type == "workflow_started":
                run_id = event.get("workflow_run_id")
                print(f"[workflow_started]  run_id={run_id}")

            elif event_type == "node_started":
                data = event.get("data", {})
                print(f"  [node_started]  {data.get('title', data.get('node_id', ''))}")

            elif event_type == "node_finished":
                data = event.get("data", {})
                outputs = data.get("outputs") or {}
                status  = data.get("status", "")
                title   = data.get("title", data.get("node_id", ""))
                print(f"  [node_finished] {title}  status={status}")
                if outputs:
                    print(f"    outputs: {json.dumps(outputs, ensure_ascii=False, indent=6)}")

            elif event_type == "workflow_finished":
                data = event.get("data", {})
                final_outputs = data.get("outputs") or {}
                status  = data.get("status", "")
                print(f"\n[workflow_finished] status={status}")
                print(f"  outputs:\n{json.dumps(final_outputs, ensure_ascii=False, indent=4)}")

            elif event_type == "text_chunk":
                chunk = event.get("data", {}).get("text", "")
                print(chunk, end="", flush=True)

            elif event_type == "error":
                print(f"[ERROR] {event}", file=sys.stderr)

            else:
                print(f"[{event_type}] {json.dumps(event, ensure_ascii=False)}")

    print(f"\n{'─'*60}\nDone.")
    return {"run_id": run_id, "outputs": final_outputs}


def _clean(value):
    """Remove characters illegal in Excel cells (ASCII control chars except tab/LF/CR)."""
    if not isinstance(value, str):
        return value
    import re
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)


def save_to_excel(file_path: str, result: dict, output_xlsx: str) -> None:
    """Append a row (file_path + workflow outputs) to an Excel file."""
    abs_path = os.path.abspath(file_path)
    outputs  = result.get("outputs", {})
    run_id   = result.get("run_id", "")
    now      = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(output_xlsx):
        wb = openpyxl.load_workbook(output_xlsx)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        header = ["timestamp", "file_path", "run_id"]
        header += [str(k) for k in outputs.keys()]
        ws.append(header)

    header_row = [cell.value for cell in ws[1]]

    for key in outputs.keys():
        col_name = str(key)
        if col_name not in header_row:
            ws.cell(row=1, column=len(header_row) + 1, value=col_name)
            header_row.append(col_name)

    row_data = [""] * len(header_row)
    fixed = {"timestamp": now, "file_path": abs_path, "run_id": run_id}
    for key, val in fixed.items():
        if key in header_row:
            row_data[header_row.index(key)] = _clean(val)
    for key, val in outputs.items():
        col_name = str(key)
        if col_name in header_row:
            cell_val = (
                val if isinstance(val, (str, int, float, bool)) or val is None
                else json.dumps(val, ensure_ascii=False)
            )
            row_data[header_row.index(col_name)] = _clean(cell_val)

    ws.append(row_data)
    wb.save(output_xlsx)
    print(f"\n[Excel] Row saved → {output_xlsx}")


def main():
    parser = argparse.ArgumentParser(description="Run a Dify workflow with a local file.")
    parser.add_argument(
        "--file", "-f",
        default=DEFAULT_FILE,
        help=f"Path to the file to upload (default: {DEFAULT_FILE})",
    )
    parser.add_argument(
        "--type", "-t",
        default="document",
        help="Dify document type hint: document / image / audio / video (default: document)",
    )
    parser.add_argument(
        "--var", "-v",
        default=FILE_VAR,
        help=f"Workflow input variable name for the file (default: {FILE_VAR})",
    )
    parser.add_argument(
        "--output", "-o",
        default=DEFAULT_OUTPUT,
        help=f"Excel file path to save results (default: {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--debug", "-d",
        action="store_true",
        help="Print the request payload before sending",
    )
    args = parser.parse_args()

    try:
        upload_result = upload_file(args.file)
        result = run_workflow(
            upload_result["id"],
            doc_type=args.type,
            file_var=args.var,
            debug=args.debug,
        )
        save_to_excel(args.file, result, args.output)
    except requests.HTTPError as e:
        print(f"\n[HTTP Error] {e}", file=sys.stderr)
        sys.exit(1)
    except FileNotFoundError:
        print(f"\n[Error] File not found: {args.file}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
